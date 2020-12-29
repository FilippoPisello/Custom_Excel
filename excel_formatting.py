# Created by Filippo Pisello
from os import path, getcwd
from typing import List, Union
import sys

import pandas as pd
from openpyxl import load_workbook

from excel_style import ExcelStyle
sys.path.append(r"C:\Users\Filippo Pisello\Desktop\Python\Git Projects\Git_Spreadsheet")
from spreadsheet import Spreadsheet

class CustomExcel(Spreadsheet):
    """
    Class to export a pandas data frame to Excel with some predefined formatting.

    ---------------
    The class intakes as the main argument a pandas dataframe. It automatically
    identifies the parts of the table: header, index, body. Three predefined
    formatting styles can be applied to each of the parts: "strong", "light",
    "main".

    These formatting styles can be modified: type help(obj.strong_formatting)
    for more. New styles can be created as well through the new_style() method.

    The main method of the class is to_custom_excel(), which allows the user
    to export the data frame to excel. It can be considered an enhanced version
    of pandas.dataframe.to_excel().

    Arguments
    ----------------
    dataframe : pandas dataframe object (mandatory)
        Dataframe to be considered.
    file_name: str (mandatory)
        The name of the file to be exported. It should be in the form "foo.xlsx".
    keep_index : Bool, default=False
        If True, it is taken into account that the first column of the spreadsheet
        will be occupied by the index.
    skip_rows: int, default=0
        The number of rows which should be left empty at the top of the spreadsheet.
    skip_cols: int, default=0
        The number of columns which should be left empty at the left of the spreadsheet.
    correct_lists: Bool, default=False
        If True, the lists stored as the dataframe entries are modified to be more
        readable in the traditional spreadsheet softwares. More details on dedicated
        docstring.
    sheet_name: str, default="Sheet1"
        The label of the sheet to be created within the Excel workbook
    header_style: str or ExcelStyle object, default="strong"
        The style to be given to the table's header. If str it should be one of
        the following keywords: "strong", "light", "plain". Custom ExcelStyle obj
        can be created through the new_style() method.
    index_style: str or ExcelStyle object, default="light"
        The style to be given to the table's index. If str it should be one of
        the following keywords: "strong", "light", "plain". Custom ExcelStyle obj
        can be created through the new_style() method.
    body_style: str or ExcelStyle object, default="plain"
        The style to be given to the table's body. If str it should be one of
        the following keywords: "strong", "light", "plain". Custom ExcelStyle obj
        can be created through the new_style() method.
    """
    # Preset formatting styles
    strong_formatting = ExcelStyle("0066cc", "ffffff", 12, True, "center")
    light_formatting = ExcelStyle("b2beb5", alignment="left")
    plain_formatting = ExcelStyle()

    def __init__(self, dataframe: pd.DataFrame, file_name, keep_index: bool=False,
                 skip_rows: int=0, skip_columns: int=0, correct_lists: bool=False,
                 sheet_name="Sheet1", header_style="strong", index_style="light",
                 body_style="plain"):
        super().__init__(dataframe, keep_index, skip_rows, skip_columns, correct_lists)
        self.file_name = file_name
        self.sheet_name = sheet_name

        self.workbook, self.sheet = None, None

        # Styles for the table parts
        self.header_style = self._style_keyword_to_obj(header_style)
        self.index_style = self._style_keyword_to_obj(index_style)
        self.body_style = self._style_keyword_to_obj(body_style)

    # -------------------------------------------------------------------------
    # 1) Main methods
    # -------------------------------------------------------------------------
    def to_custom_excel(self, custom_width=20, check_file_name=True):
        """
        Exports pandas dataframe to Excel with some formatting.

        ---------------
        First, the pandas dataframe is saved to Excel with the given file and
        sheet name. Three scenarios for this process: (1) if no workbook under the
        given name exists, it gets created. (2) If the workbook exists but there
        is no sheet with given name, a new sheet gets appended. (3) If both
        workbook and sheet exists, the latter gets overwritten.

        Then the table gets formatted: header, index and body can have different
        styles. These are chosen through the object attributes obj.header_style,
        obj.index_style and obj.body_style. Type help(CustomExcel) for more.

        Arguments
        ----------------
        custom_width: int, default=20
        - Width to be attributed to all the columns of the spreadsheet, expressed
          in points.
        check_file_name: Bool, default=True
        - If True, the program will try to add the ".xlsx" file extension at the
        end of the file name if it is not present, to avoid errors. If False no
        control is carried out.
        """
        # File saving process
        if check_file_name:
            self._correct_file_name()
        self.save_df_to_excel()


        # Table customization
        self.workbook, self.sheet = self._get_workbook_sheet()
        #Loop to apply styles to the different table's components
        components = [self.body, self.header, self.index]
        styles = [self.body_style, self.header_style, self.index_style]
        for component, style in zip(components, styles):
            self.format_cells(component, style)
        # Setting columns' width
        self._adjust_all_columns_width(custom_width)

        # Save edits
        self.workbook.save(filename=self.file_name)
        return

    @staticmethod
    def new_style(fill_color=None, font_color="000000", font_size=11,
                  font_bold=False, alignment="center"):
        """
        Returns an ExcelStyle object

        Arguments
        ----------------
        fill_color: str, default=None
        - Fill color of the cells. If None, no fill color is applied.
        font_color: str, default="000000"
        - Font color of the cells. Default color is black.
        font_size: int, default=11
        - Size of the cell font.
        font_bold: Bool, default=False
        - If True cell text is bold.
        alignment: str, default="center"
        - Horizontal alignment of the text content. It can be either "center",
        "right" or "left".
        """
        style = ExcelStyle(fill_color, font_color, font_size, font_bold, alignment)
        return style

    # -------------------------------------------------------------------------
    # 2 - Worker Methods
    # -------------------------------------------------------------------------
    # 2.1 - Methods used in attributes
    # --------------------------------
    def _style_keyword_to_obj(self, style_input: Union[str, ExcelStyle]):
        """
        Returns input if input is ExcelStyle obj, otherwise returns the ExcelStyle
        obj associated with the given keyword

        ---------------
        The function is meant to convert the user input for style into an
        appropriate style object later to be used throughout the function.

        There are three built in ExcelStyle objects which can accessed through
        keywords. These are the keys of name_style_dict.
        """
        if isinstance(style_input, ExcelStyle):
            return style_input
        # Style name to style object
        name_style_dict = {"strong" : self.strong_formatting,
                           "light" : self.light_formatting,
                           "plain" : self.plain_formatting}
        try:
            return name_style_dict[style_input]
        except KeyError as e:
            txt = list(name_style_dict.keys())
            msg = f"Provide ExcelStyle object or keyword among the following: {txt}"
            raise KeyError(msg) from e

    # --------------------------------
    # 2.2 - Chain of methods used in to_custom_excel for the file saving process
    # --------------------------------
    def _correct_file_name(self, required_extension=".xlsx"):
        """
        Adds desired extention if the user did not include it in file name.

        ---------------
        Example:
        - File name without file extension: "Foo" --> "Foo.xlsx"
        """
        if not self.file_name.endswith(required_extension):
            self.file_name = self.file_name + required_extension
            print(f"The file name provided was modified to '{self.file_name}' "
                  "to avoid errors in the program execution. To deactivate this "
                  "autocorrection turn to 'False' the default arg 'check_file_name'.")
        return

    def save_df_to_excel(self):
        """
        Saves pandas dataframe to excel. Creates new workbook or appends to old ones.

        ---------------
        Three outcomes:
        - If no workbook exists under the provided name:
            1) A new workbook is created
        - If it exists and if it already exists a sheet with the name given:
            2) The sheet gets overwritten
        - If it exists and no sheet with the name given exists:
            3) A new sheet is appended to the workbook
        """
        workbook, sheet = self._check_file_existence()

        if workbook is None:
            self.df.to_excel(self.file_name, sheet_name=self.sheet_name,
                            index=self.keep_index, startrow=self.skip_rows,
                            startcol=self.skip_cols)
        else:
            if sheet is not None: # it needs to be removed first
                workbook.remove_sheet(sheet)
                workbook.save(self.file_name)
            with pd.ExcelWriter(self.file_name, mode="a") as writer:
                self.df.to_excel(writer, sheet_name=self.sheet_name,
                                index=self.keep_index, startrow=self.skip_rows,
                                startcol=self.skip_cols)
        return

    def _check_file_existence(self):
        """
        Returns two items: first provides info on the existence of workbook called
        as self.file_name and second on sheet called as self.sheet_name.

        ---------------
        It returns two items:
        - Workbook/None: workbook object if one called as the provided file name
        exists, else None.
        - Sheet/None: sheet object if workbook contains a sheet with the given
        name, else None.

        Note that if the following three conditions apply: (1) workbook exists,
        (2) sheet exists, (3) workbook has only one sheet, then None, None it is
        returned. This is because in the saving process, replacing a single sheet
        is equivalent to overwriting the whole workbook.
        """
        if path.exists(getcwd() + "/" + self.file_name):
            workbook = load_workbook(self.file_name)
            try:
                sheet = workbook[self.sheet_name]
            except KeyError:
                # Case: workbook exists, sheet doesn't exist
                return workbook, None
            if len(workbook.sheetnames) > 1:
                # Case: workbook exists, sheet exists and it's not unique
                return workbook, sheet

        # Cases: (1) workbook doesn't exists (2) sheet exists but it's unique
        return None, None

    # --------------------------------
    # 2.3 - Chain of methods used in to_custom_excel for formatting
    # --------------------------------
    def _get_workbook_sheet(self):
        """
        Returns the desired workbook and sheet objects based on their name.
        """
        workbook = load_workbook(self.file_name)
        sheet = workbook.get_sheet_by_name(self.sheet_name)
        return workbook, sheet

    def _adjust_all_columns_width(self, custom_width):
        """
        Sets the width of all the columns of the sheet to a fixed value
        """
        for value in range(self.index_coordinates[0][0], self.body_coordinates[1][0] + 1):
            self.sheet.column_dimensions[self.letter_from_index(value)].width = custom_width
        return

    def format_cells(self, cells_list: List[str], style_object: ExcelStyle):
        """
        Applies a given formatting style to a target set of cells

        ---------------
        Cells_list should be a list in the form ["A1", "A2", ...].
        """
        for cell in cells_list:
            self.sheet[cell].font = style_object.font()
            self.sheet[cell].fill = style_object.fill()
            self.sheet[cell].alignment = style_object.alignment()
        return
